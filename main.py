#libraries
from win32com.client import Dispatch
from tkinter import *
from tkinter import messagebox
import os
import json
import pyodbc as pyo
from openpyxl import Workbook, load_workbook
from tkinter import ttk
import shutil
import os.path,time
import sqlite3
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
#code
root = Tk()
def get_size(start_path = '.'):
    total_size = 0
    for dirpath, dirnames, filenames in os.walk(start_path):
        for f in filenames:
            fp = os.path.join(dirpath, f)
            # skip if it is symbolic link
            if not os.path.islink(fp):
                total_size += os.path.getsize(fp)

    return total_size


def close(win):
    win.destroy()


NAV_TOP = 0
def usb(usb_type_letter):
    usbconnect = Tk()
    try:
        time.sleep(2)

        Label(usbconnect, text='Name: ' + usb_type_letter).place(x=20, y=20)
        print(usb_type_letter)
        l = os.listdir(usb_type_letter)

        print(l)
        x = 100
        y = 100
        num = 0
        main_frame = Frame(usbconnect)
        main_frame.pack(fill=BOTH, expand=1)

        # Create A Canvas
        my_canvas = Canvas(main_frame)
        my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

        # Add A Scrollbar To The Canvas
        my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
        my_scrollbar.pack(side=RIGHT, fill=Y)

        # Configure The Canvas
        my_canvas.configure(yscrollcommand=my_scrollbar.set)
        my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))

        # Create ANOTHER Frame INSIDE the Canvas
        second_frame = Frame(my_canvas)

        # Add that New frame To a Window In The Canvas
        my_canvas.create_window((0, 0), window=second_frame, anchor="nw")
        v = 1


        def delete(name):
            try:
                os.rmdir(usb_type_letter + name)


            except PermissionError as md:
                messagebox.showinfo("USB VIEWER - Access denied", md)
            except NotADirectoryError as d:
                print(d)
                os.remove(usb_type_letter + name)
            except OSError:
                shutil.rmtree(usb_type_letter + name)
            finally:
                usbconnect.destroy()
                usb(usb_type_letter)

        def openfile(name):
            file_dr = usb_type_letter + name + "/"
            if os.path.isdir(file_dr):
                usbconnect.destroy()
                usb(file_dr)
            else:
                print(file_dr)
                os.startfile(file_dr)

        class class_btn:

            def __init__(self, name_btn):
                self.name_btn = name_btn

                def option(name):
                    file_dr = usb_type_letter + name
                    print(name)
                    global NAV_TOP
                    #Button(NAV_TOP, text="Open file", command=lambda: openfile(name)).place(x=300, y=0)
                    #Button(NAV_TOP, text="Delete file", command=lambda: delete(name)).place(x=400, y=0)
                    NAV_TOP.add_cascade(label="Open file", command=lambda:openfile(name))
                    NAV_TOP.add_cascade(label="Delete file", command=lambda: delete(name))
                    dt = Frame(usbconnect, width=500, height=usbconnect.winfo_screenheight())
                    dt.place(x=510, y=20)
                    Label(dt, text="Name: " + name).place(x=10, y=10)
                    if os.path.isdir(file_dr):
                        file = "folder"
                        size = str(get_size(file_dr))

                    else:
                        file_ = name.split(".")
                        file = repr(file_[-1])
                        size = str(os.path.getsize(file_dr))

                    Label(dt, text="Directory: " + file_dr).place(x=10, y=30)
                    Label(dt, text="Type: " + file).place(x=10, y=50)
                    Label(dt, text="Date created: " + time.ctime(os.path.getctime(file_dr))).place(x=10, y=70)
                    Label(dt, text="Last modification: " + time.ctime(os.path.getmtime(file_dr))).place(x=10, y=90)

                    Label(dt, text="Size: " + size + " BYTES").place(x=10, y=120)




                print(self.name_btn)
                bt = Button(second_frame, text=self.name_btn, command=lambda: option(self.name_btn))
                bt.grid(row=y, column=x, pady=20, padx=10)

        for i in l:
            widthscreen = 2

            if num == widthscreen:
                num = 0
                y = y + 100
                x = 100
                print(y)
                print(x)
                print(num)
            bt_name = i
            class_btn(i)

            num = num + 1
            x = x + 100
            v = v + 1
            print(num)
        template(usbconnect)
        usbconnect.title('USB VIEWER - ' + usb_type_letter)
        def export_data():

            ms = messagebox.askyesno("USB VIEWER", "Are you sure?")
            print(ms)

            if ms:
                def doesFileExists(filePathAndName):
                    return os.path.exists(filePathAndName)
                if doesFileExists('c:/export'):
                    pass
                else:
                    os.chdir("C:/")
                    os.makedirs("Export")

                con = sqlite3.connect("C:/export/data.db")
                cursor = con.cursor()
                cursor.execute(
                    " CREATE TABLE IF NOT EXISTS  Disk (Name TEXT , Type TEXT , Directory TEXT , Date_Created TEXT, Last_Modification TEXT, Size TEXT)")
                for i in l:
                    file_dr = usb_type_letter + i
                    if os.path.isdir(file_dr):
                        file = "folder"
                        size = get_size(file_dr)

                    else:
                        file_ = i.split(".")
                        file = repr(file_[-1])
                        size = str(os.path.getsize(file_dr))

                    co = sqlite3.connect("C:/export/data.db")
                    cursor = co.cursor()
                    cursor.execute('INSERT INTO Disk VALUES(?, ?, ?, ?, ?, ?)', (i, file, file_dr, time.ctime(os.path.getctime(file_dr)), time.ctime(os.path.getmtime(file_dr)), size))
                    co.commit()
                os.startfile("C:/export/")
                co.close()


        def PDF():
            messagebox.showinfo("USB - VIEWER", "If you have any undifined character they will be  saved as black squares.")
            connect = sqlite3.connect("data.db")
            cursor = connect.cursor()
            cursor.execute(
                " CREATE TABLE IF NOT EXISTS  Disk (Name TEXT , Type TEXT , Directory TEXT , Date_Created TEXT, Last_Modification TEXT, Size TEXT)")
            cursor.execute('INSERT INTO Disk VALUES(?, ?, ?, ?, ?, ?)',
                           ("Name", "Type", "Directory", "Date Created", "Last Modification", "Size"))
            connect.commit()
            for i in l:
                file_dr = usb_type_letter + i
                if os.path.isdir(file_dr):
                    file = "folder"
                    size = get_size(file_dr)

                else:
                    file_ = i.split(".")
                    file = repr(file_[-1])
                    size = str(os.path.getsize(file_dr))

                cursor.execute('INSERT INTO Disk VALUES(?, ?, ?, ?, ?, ?)', (
                    i, file, file_dr, time.ctime(os.path.getctime(file_dr)), time.ctime(os.path.getmtime(file_dr)),
                    size))
                connect.commit()
            if os.access("Export.pdf", 0):
                os.remove("Export.pdf")
            doc = SimpleDocTemplate("Export.pdf", pagesize=landscape(letter))

            data = (cursor.execute("SELECT * FROM Disk").fetchall())
            print(data)
            table = Table(data)
            table.setStyle(TableStyle([

                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            # Add the table to the PDF document
            doc.build([table])
            time.sleep(2)
            connect.close()
            os.startfile("Export.pdf")
            os.remove("data.db")
        def Excel():
            if os.access("Export.xlsx", 0):
                os.remove("Export.xlsx")
                wb = Workbook()
                ws = wb.active
                ws.title = "Export"
                wb.save("Export.xlsx")
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Export"
                wb.save("Export.xlsx")
            wb = load_workbook("Export.xlsx")
            sheet = wb.active
            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Type"
            sheet.cell(row=1, column=3).value = "Directory"
            sheet.cell(row=1, column=4).value = "Date Created"
            sheet.cell(row=1, column=5).value = "Last Modification"
            sheet.cell(row=1, column=6).value = "Size"
            for i in l:
                file_dr = usb_type_letter + i
                if os.path.isdir(file_dr):
                    file = "folder"
                    size = get_size(file_dr)

                else:
                    file_ = i.split(".")
                    file = repr(file_[-1])
                    size = str(os.path.getsize(file_dr))
                max = sheet.max_row
                sheet.cell(column=1, row=max + 1, value=i)
                sheet.cell(column=2, row=max + 1, value=file)
                sheet.cell(column=3, row=max + 1, value=file_dr)
                sheet.cell(column=4, row=max + 1, value=time.ctime(os.path.getctime(file_dr)))
                sheet.cell(column=5, row=max + 1, value= time.ctime(os.path.getmtime(file_dr)))
                sheet.cell(column=6, row=max + 1, value=size)

            wb.save("Export.xlsx")
            os.startfile("Export.xlsx")

        def JSON():
            list_d = []
            for i in l:
                file_dr = usb_type_letter + i
                if os.path.isdir(file_dr):
                    file = "folder"
                    size = get_size(file_dr)

                else:
                    file_ = i.split(".")
                    file = repr(file_[-1])
                    size = str(os.path.getsize(file_dr))
                c = "Name: " + i + " Type: " + file + " Directory: " + file_dr + " Date created: " + time.ctime(os.path.getctime(file_dr)) + " Last modification: " + time.ctime(os.path.getmtime(file_dr)) + "Size: " + str(size)
                list_d.append(c)
            with open('data.json', 'w') as f:
                json.dump(list_d, f, indent=2)
                os.startfile('data.json')

        global NAV_TOP

        export_menu = Menu(NAV_TOP)
        dev = Menu(export_menu)
        NAV_TOP.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export as PDF", command=PDF)
        export_menu.add_command(label="Export as Excel", command=Excel)
        export_menu.add_separator()
        export_menu.add_cascade(label="Developer options", menu=dev)
        dev.add_command(label="Export as database", command=export_data)
        dev.add_command(label="Export to JSON", command=JSON)
       

    except PermissionError as m:

        usbconnect.destroy()
        messagebox.showinfo('USB VIEWER - Access denied', m)

def connect_usb():
    root.destroy()
    usb_connect = Tk()
    def usbnext():
        usb_type_letter = usb_type.get() + ":/"

        def doesFileExists(usb):
            return os.path.exists(usb)

        if doesFileExists(usb_type_letter):

            usb_connect.destroy()
            usb(usb_type_letter)
        else:
            print('false')
            messagebox.showerror('ERROR', 'Please try again later')



    template(usb_connect)
    leb = Label(usb_connect, text='USB LETTER').place(x=200, y=150)
    usb_type = Entry(usb_connect)
    usb_type.place(x=200, y=200)
    btn = Button(usb_connect, text='Next', command=usbnext).place(x=200, y=250)




def template(win):
    win.title('USB VIEWER')
    win.geometry('500x500')
    widthscreen = win.winfo_screenwidth()
    global NAV_TOP
    print(widthscreen)
    NAV_TOP = Menu(win)
    win.config(menu=NAV_TOP)
    CON = Menu(NAV_TOP)
    NAV_TOP.add_cascade(label="connect to usb", command=connect_usb)
    NAV_TOP.add_cascade(label="Exit", command=lambda:close(win))



template(root)
root.mainloop()